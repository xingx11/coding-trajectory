"""Project pool management for cross-person deduplication.

Provides commands to pre-search GitHub/Gitee candidates, assign them
to persons, and query pool status.
"""

from __future__ import annotations

import json
import math
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from ctpipe.github_search import (
    GitHubRepo,
    search_projects,
    search_projects_gitee,
)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

POOL_FILE = "project_pool.json"
ASSIGNMENTS_FILE = "pool_assignments.json"

# xlsx column indices (0-based) in the first sheet, starting from data row 5
_COL_TASK_TYPE = 1
_COL_DOMAIN = 2
_COL_LANGUAGE = 3
_COL_WEIGHT = 4
_COL_COUNT = 9  # "项目数量" column


# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------

def _read_weights_xlsx(path: Path) -> list[dict]:
    """Read (task_type, domain, language, weight, count) rows from xlsx."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for reading xlsx files. "
            "Install with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        task_type = str(row[_COL_TASK_TYPE] or "")
        domain = str(row[_COL_DOMAIN] or "")
        language = str(row[_COL_LANGUAGE] or "")
        weight = float(row[_COL_WEIGHT] or 0)
        count = int(row[_COL_COUNT] or 0)
        if task_type and domain and language:
            rows.append({
                "task_type": task_type,
                "domain": domain,
                "language": language,
                "weight": weight,
                "count": count,
            })
    return rows


def _compute_pool_needs(
    rows: list[dict],
    per_project: int,
    buffer: float = 1.5,
) -> dict[str, int]:
    """Compute number of repos needed per (domain, language) pair.

    Groups by (domain, language), sums the data count, divides by
    per_project, then adds a buffer multiplier.

    Returns dict like ``{"backend_service:java": 380, ...}``.
    """
    dl_counts: dict[str, int] = defaultdict(int)
    for r in rows:
        key = f"{r['domain']}:{r['language']}"
        dl_counts[key] += r["count"]

    needs: dict[str, int] = {}
    for key, total in dl_counts.items():
        raw = math.ceil(total / per_project)
        needs[key] = math.ceil(raw * buffer)
    return needs


# ---------------------------------------------------------------------------
# Pool generation (bulk search)
# ---------------------------------------------------------------------------

def _search_bulk_github(
    domain: str,
    language: str,
    needed: int,
    github_token: str = "",
    http_proxy: str = "",
    existing_repos: set[str] | None = None,
) -> list[dict]:
    """Search GitHub with multiple strategies to collect many candidates.

    Uses star-range segmentation and multiple sort strategies to exceed
    the single-query 1000-result limit.
    """
    all_repos: dict[str, GitHubRepo] = {}  # full_name -> repo (dedup)
    if existing_repos:
        # Pre-populate exclusion set so we don't re-fetch known repos
        for name in existing_repos:
            all_repos[name] = None  # type: ignore[assignment]

    star_ranges = [
        (20, 500),
        (500, 2000),
        (2000, 10000),
    ]
    sort_strategies = [
        ("stars", "desc"),
        ("stars", "asc"),
        ("updated", "desc"),
        ("", "desc"),  # best-match
    ]

    for min_s, max_s in star_ranges:
        for sort_val, order_val in sort_strategies:
            if len(all_repos) >= needed + len(existing_repos or set()):
                break
            remaining = needed - (len(all_repos) - len(existing_repos or set()))
            if remaining <= 0:
                break
            exclude = set(all_repos.keys())
            try:
                repos = search_projects(
                    domain, language,
                    count=min(remaining, 100),
                    min_stars=min_s,
                    max_stars=max_s,
                    sort=sort_val,
                    order=order_val,
                    exclude_repos=exclude,
                    github_token=github_token,
                    http_proxy=http_proxy,
                )
                for r in repos:
                    if r.full_name not in all_repos:
                        all_repos[r.full_name] = r
            except Exception as exc:
                print(f"    WARNING: search failed ({sort_val}/{order_val} stars:{min_s}..{max_s}): {exc}")
            time.sleep(1)  # rate limit courtesy
        if len(all_repos) - len(existing_repos or set()) >= needed:
            break

    # Remove placeholder entries (existing repos) and build output
    return [
        {
            "full_name": r.full_name,
            "clone_url": r.clone_url,
            "description": r.description,
            "stars": r.stars,
        }
        for r in all_repos.values()
        if r is not None
    ]


def _search_bulk_gitee(
    domain: str,
    language: str,
    needed: int,
    gitee_token: str = "",
) -> list[dict]:
    """Search Gitee for bulk candidates."""
    all_repos: dict[str, GitHubRepo] = {}

    exclude = set()
    try:
        repos = search_projects_gitee(
            domain, language,
            count=min(needed, 50),
            min_stars=10,
            exclude_repos=exclude,
            gitee_token=gitee_token,
        )
        for r in repos:
            if r.full_name not in all_repos:
                all_repos[r.full_name] = r
    except Exception as exc:
        print(f"    WARNING: gitee search failed: {exc}")

    return [
        {
            "full_name": r.full_name,
            "clone_url": r.clone_url,
            "description": r.description,
            "stars": r.stars,
        }
        for r in all_repos.values()
    ]


def pool_generate(
    base_dir: Path,
    weights_path: Path,
    per_project: int = 5,
    source: str = "github",
    github_token: str = "",
    gitee_token: str = "",
    http_proxy: str = "",
    buffer: float = 1.5,
) -> int:
    """Generate project pool from xlsx weights.

    Searches GitHub/Gitee for candidate repos per (domain, language) pair.
    Writes ``project_pool.json`` to *base_dir*.

    If a pool file already exists, runs in **incremental mode**: keeps
    existing repos and only searches for keys that have shortfalls.
    """
    print(f"Reading weights from {weights_path}...")
    rows = _read_weights_xlsx(weights_path)
    print(f"  {len(rows)} scenario rows loaded")

    needs = _compute_pool_needs(rows, per_project, buffer)
    total_needed = sum(needs.values())
    print(f"  {len(needs)} (domain, language) pairs, {total_needed} repos needed (with {buffer}x buffer)")

    # Load existing pool for incremental mode
    pool_path = base_dir / POOL_FILE
    existing_pool: dict[str, dict] = {}
    if pool_path.exists():
        try:
            old_data = json.loads(pool_path.read_text(encoding="utf-8"))
            existing_pool = old_data.get("pools", {})
            existing_total = sum(len(p.get("repos", [])) for p in existing_pool.values())
            print(f"  Incremental mode: existing pool has {existing_total} repos")
        except Exception as exc:
            print(f"  WARNING: failed to read existing pool ({exc}), starting fresh")
    print()

    pool: dict[str, dict] = {}
    total_found = 0
    skipped = 0
    shortfalls: list[str] = []

    for i, (key, needed) in enumerate(sorted(needs.items(), key=lambda x: -x[1]), 1):
        domain, language = key.split(":", 1)

        # Check existing pool for this key
        existing_repos_list = existing_pool.get(key, {}).get("repos", [])
        existing_count = len(existing_repos_list)
        existing_names = {r["full_name"] for r in existing_repos_list}

        gap = needed - existing_count
        if gap <= 0:
            # Already have enough
            pool[key] = {"needed": needed, "repos": existing_repos_list}
            total_found += existing_count
            skipped += 1
            continue

        print(f"[{i}/{len(needs)}] {key}: need {needed}, have {existing_count}, searching {gap} more...")

        if source == "gitee":
            new_repos = _search_bulk_gitee(domain, language, gap, gitee_token)
        else:
            new_repos = _search_bulk_github(
                domain, language, gap,
                github_token, http_proxy,
                existing_repos=existing_names,
            )

        # Merge: existing + new (dedup)
        merged = list(existing_repos_list)
        seen = set(existing_names)
        for r in new_repos:
            if r["full_name"] not in seen:
                merged.append(r)
                seen.add(r["full_name"])

        pool[key] = {"needed": needed, "repos": merged}
        found = len(merged)
        total_found += found

        if found < needed:
            shortfalls.append(f"  {key}: need {needed}, found {found} (shortfall: {needed - found})")
            print(f"  WARNING: only found {found}/{needed} (was {existing_count}, added {found - existing_count})")
        else:
            print(f"  OK: {found} repos (was {existing_count}, added {found - existing_count})")

        # Rate limit between searches
        if i < len(needs):
            time.sleep(2)

    if skipped:
        print(f"\n  Skipped {skipped} keys (already sufficient)")

    # Write pool file
    output = {
        "_generated": datetime.now().isoformat(timespec="seconds"),
        "_source": source,
        "_total_repos": total_found,
        "_per_project": per_project,
        "_buffer": buffer,
        "pools": pool,
    }
    tmp = pool_path.with_suffix(".tmp")
    tmp.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp.replace(pool_path)

    print(f"\nPool generated: {pool_path}")
    print(f"  Total repos: {total_found}/{total_needed}")

    if shortfalls:
        print(f"\n  Shortfalls ({len(shortfalls)}):")
        for s in shortfalls:
            print(s)

    return 0


# ---------------------------------------------------------------------------
# Pool assignment
# ---------------------------------------------------------------------------

def pool_assign(
    base_dir: Path,
    persons: int,
    per_project: int = 5,
    weights_path: Path | None = None,
    tasks_per_person: int = 45,
) -> int:
    """Assign pool repos to persons, optionally weighted by scenario counts.

    When *weights_path* is provided, each person receives repos in each
    ``domain:language`` key proportional to that key's task-count share,
    so the final data distribution matches the weight table.

    Without *weights_path*, falls back to simple round-robin (backward compat).

    **Incremental mode**: if ``pool_assignments.json`` already exists, keeps
    each person's existing repos and only appends newly available repos from
    the pool.  This ensures previously used repos stay with their original
    person across pool refreshes.

    Reads ``project_pool.json`` and writes ``pool_assignments.json``.
    """
    pool_path = base_dir / POOL_FILE
    if not pool_path.exists():
        print(f"ERROR: pool file not found: {pool_path}")
        print("  Run 'ctpipe pool generate' first.")
        return 1

    pool_data = json.loads(pool_path.read_text(encoding="utf-8"))
    pools = pool_data.get("pools", {})

    # Load existing assignments for incremental mode
    assign_path = base_dir / ASSIGNMENTS_FILE
    old_assignments: dict[str, dict[str, list[str]]] = {}
    if assign_path.exists():
        try:
            old_data = json.loads(assign_path.read_text(encoding="utf-8"))
            old_assignments = old_data.get("assignments", {})
            old_total = sum(
                sum(len(v) for v in pa.values())
                for pa in old_assignments.values()
            )
            print(f"Incremental mode: keeping {old_total} existing repo assignments")
        except Exception as exc:
            print(f"WARNING: failed to read existing assignments ({exc}), starting fresh")

    # Build assignments: person_id -> {pool_key -> [repo_full_names]}
    assignments: dict[str, dict[str, list[str]]] = {
        str(p): {} for p in range(1, persons + 1)
    }

    total_assigned = 0

    if weights_path:
        # --- Weighted proportional assignment ---
        rows = _read_weights_xlsx(weights_path)
        dl_counts: dict[str, int] = defaultdict(int)
        for r in rows:
            key = f"{r['domain']}:{r['language']}"
            dl_counts[key] += r["count"]
        total_count = sum(dl_counts.values())

        projects_per_person = math.ceil(tasks_per_person / per_project)
        buffer = 1.5

        print(f"Weighted assignment: {tasks_per_person} tasks/person, "
              f"{projects_per_person} projects/person, {buffer}x buffer")

        for key, pool_info in sorted(pools.items()):
            repos = pool_info.get("repos", [])
            all_repo_names = {r["full_name"] for r in repos}
            if not all_repo_names:
                continue

            # Compute per-person target based on weight ratio
            key_count = dl_counts.get(key, 0)
            if total_count > 0 and key_count > 0:
                ratio = key_count / total_count
                target = max(1, math.ceil(ratio * projects_per_person * buffer))
            else:
                target = 1

            # Collect what each person already has for this key
            # and build the set of all already-assigned repos
            already_assigned: set[str] = set()
            for p in range(1, persons + 1):
                pid = str(p)
                existing = old_assignments.get(pid, {}).get(key, [])
                # Keep only repos still in the pool (prune stale refs)
                kept = [r for r in existing if r in all_repo_names]
                assignments[pid][key] = kept
                already_assigned.update(kept)

            # New repos available for distribution
            new_repos = [r["full_name"] for r in repos if r["full_name"] not in already_assigned]
            new_idx = 0

            # Top up each person to reach target
            for p in range(1, persons + 1):
                pid = str(p)
                current = assignments[pid][key]
                need = target - len(current)
                if need > 0 and new_idx < len(new_repos):
                    take = new_repos[new_idx:new_idx + need]
                    current.extend(take)
                    new_idx += len(take)
                if current:
                    assignments[pid][key] = current
                    total_assigned += len(current)
                else:
                    assignments[pid].pop(key, None)

            shortfall_persons = sum(
                1 for p in range(1, persons + 1)
                if len(assignments[str(p)].get(key, [])) < target
            )
            if shortfall_persons:
                print(f"  WARNING: {key}: target {target}/person but "
                      f"{shortfall_persons} person(s) have fewer")
    else:
        # --- Legacy round-robin assignment ---
        for key, pool_info in sorted(pools.items()):
            repos = pool_info.get("repos", [])
            repo_names = [r["full_name"] for r in repos]

            for i, name in enumerate(repo_names):
                person_id = str((i % persons) + 1)
                if key not in assignments[person_id]:
                    assignments[person_id][key] = []
                assignments[person_id][key].append(name)
                total_assigned += 1

    # Write assignments file
    output = {
        "_generated": datetime.now().isoformat(timespec="seconds"),
        "_persons": persons,
        "_per_project": per_project,
        "_total_assigned": total_assigned,
        "assignments": assignments,
    }
    tmp = assign_path.with_suffix(".tmp")
    tmp.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp.replace(assign_path)

    # Print summary
    print(f"\nAssignments written: {assign_path}")
    print(f"  {persons} persons, {total_assigned} total repo assignments")
    print()

    for pid in sorted(assignments.keys(), key=int):
        person_repos = sum(len(v) for v in assignments[pid].values())
        domains = len(assignments[pid])
        print(f"  Person {pid:>3s}: {person_repos:4d} repos across {domains} domain/language pairs")

    return 0


# ---------------------------------------------------------------------------
# Pool status
# ---------------------------------------------------------------------------

def _load_all_used_repos(base_dir: Path) -> set[str]:
    """Scan all delivery_*/pipeline_state.json for used repos."""
    used = set()
    for ps in sorted(base_dir.glob("delivery_*/pipeline_state.json")):
        try:
            d = json.loads(ps.read_text(encoding="utf-8"))
            used.update(d.get("_gen_used_repos", []))
        except Exception:
            pass
    return used


def pool_status(
    base_dir: Path,
    person_id: str | None = None,
) -> int:
    """Show pool status: total, used, remaining."""
    assign_path = base_dir / ASSIGNMENTS_FILE
    if not assign_path.exists():
        print(f"ERROR: assignments file not found: {assign_path}")
        print("  Run 'ctpipe pool assign' first.")
        return 1

    data = json.loads(assign_path.read_text(encoding="utf-8"))
    assignments = data.get("assignments", {})
    used_repos = _load_all_used_repos(base_dir)

    if person_id:
        person_ids = [person_id]
        if person_id not in assignments:
            print(f"ERROR: person_id '{person_id}' not found in assignments")
            return 1
    else:
        person_ids = sorted(assignments.keys(), key=int)

    print(f"Pool Status (used repos across all deliveries: {len(used_repos)})")
    print(f"{'Person':>8s}  {'Total':>6s}  {'Used':>6s}  {'Remain':>6s}  {'Usage%':>6s}")
    print("-" * 40)

    for pid in person_ids:
        person_data = assignments[pid]
        total = sum(len(v) for v in person_data.values())
        person_repos = set()
        for repos in person_data.values():
            person_repos.update(repos)
        used = len(person_repos & used_repos)
        remain = total - used
        pct = f"{used / total * 100:.0f}%" if total else "N/A"
        print(f"{pid:>8s}  {total:>6d}  {used:>6d}  {remain:>6d}  {pct:>6s}")

        if person_id:
            # Show per-domain breakdown for single person
            print()
            print(f"  {'Domain:Language':<35s}  {'Total':>5s}  {'Used':>5s}  {'Remain':>5s}")
            print(f"  {'-' * 55}")
            for key in sorted(person_data.keys()):
                repos = person_data[key]
                key_used = len(set(repos) & used_repos)
                key_remain = len(repos) - key_used
                print(f"  {key:<35s}  {len(repos):>5d}  {key_used:>5d}  {key_remain:>5d}")

    return 0


# ---------------------------------------------------------------------------
# Integration helper (used by gen.py)
# ---------------------------------------------------------------------------

def load_pool_assignments(
    base_dir: Path,
    person_id: str,
) -> dict[str, list[str]] | None:
    """Load pool assignments for a specific person.

    Returns dict ``{domain:language -> [repo_full_name, ...]}`` or None
    if no assignments file exists or person_id not found.
    """
    assign_path = base_dir / ASSIGNMENTS_FILE
    if not assign_path.exists():
        return None

    try:
        data = json.loads(assign_path.read_text(encoding="utf-8"))
        assignments = data.get("assignments", {})
        person_data = assignments.get(person_id)
        if person_data is None:
            return None
        return person_data
    except Exception:
        return None
